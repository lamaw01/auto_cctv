from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import os
from time import ctime
from openpyxl import Workbook
from openpyxl import load_workbook

timeout = 5

#web elements
login_button_element = '//*[@id="login"]/table/tbody/tr/td[2]/div/div[5]/button'
cam_config_element = '//*[@id="nav"]/li[5]/a'
maintenance_element = '//*[@id="menu"]/div/div[2]/div[3]/span'
reboot_button_element = '//*[@id="maintainUpgrade"]/div[1]/div[2]/span[1]/button'
reboot_ok_button_element = '//*[@id="config"]/div[1]/div/table/tbody/tr[2]/td[2]/div/table/tbody/tr[3]/td/div/button[1]'

#open ip cam page
def reboot(ip,name,x,y,nvr):
   print('accessing...',name)
   try:
      chrome_options = Options()
      #chrome_options.add_argument("--incognito")
      chrome_options.add_experimental_option("detach", True)
      cam_driver = webdriver.Chrome(options=chrome_options)
      #set position and size
      cam_driver.set_window_size(1280, 720)
      cam_driver.set_window_position(x, y, windowHandle='current')
      cam_driver.get('http://'+ip)
      success = True
      #credentials of selected ip cam
      username = 'admin'
      password1 = '123456@ad'
      password2 = 'scores135792468'
      time.sleep(timeout)
      wait_for_element_load(login_button_element,cam_driver)
      pre_login_url = cam_driver.current_url
      #login using first password
      cam_driver.find_element(By.ID,'username').send_keys(username)
      cam_driver.find_element(By.ID,'password').send_keys(password1)
      cam_driver.find_element(By.XPATH,login_button_element).click()
      time.sleep(timeout)
      post_login_url = cam_driver.current_url
      #check if password is error
      if pre_login_url == post_login_url:
         cam_driver.find_element(By.ID,'username').send_keys(username)
         cam_driver.find_element(By.ID,'password').send_keys(password2)
         cam_driver.find_element(By.XPATH,login_button_element).click()
         time.sleep(timeout)
         wait_for_element_load(cam_config_element,cam_driver)
         cam_driver.find_element(By.XPATH, cam_config_element).click()
         wait_for_element_load(maintenance_element,cam_driver)
         cam_driver.find_element(By.XPATH,maintenance_element).click()
         wait_for_element_load(reboot_button_element,cam_driver)
         cam_driver.find_element(By.XPATH,reboot_button_element).click()
         wait_for_element_load(reboot_ok_button_element,cam_driver)
         cam_driver.find_element(By.XPATH,reboot_ok_button_element).click()
         time.sleep(15)
   except Exception as e:
      print('reboot camera error',e)
      success = False
   finally:
      #close current browser
      cam_driver.close()
      #write log
      write_excel(ip,name,success,nvr)
      print('is rebooted?..'+name,success)

#function to ensure web element loaded
def wait_for_element_load(element, driver):
   element_wait_load = EC.presence_of_element_located((By.XPATH, element))
   WebDriverWait(driver, 10).until(element_wait_load)
   time.sleep(timeout)

#write logs
def write_excel(ip,name,status,nvr):
   file_name = 'logs.xlsx'
   new_row = [ctime(),ip,name,status,nvr]
   is_file_exist = os.path.isfile(file_name)
   if is_file_exist:
      wb = load_workbook(file_name)
      ws = wb.active
      ws.append(new_row)
   else:
      wb = Workbook()
      ws = wb.active
      ws.append(['Date', 'Ip', 'Name','Rebooted','NVR'])
      ws.append(new_row)
   wb.save(file_name)