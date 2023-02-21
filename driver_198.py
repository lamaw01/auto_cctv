from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import os
from time import ctime
from openpyxl import Workbook
from openpyxl import load_workbook

#web elements
login_button_element = '//*[@id="login"]/table/tbody/tr/td[2]/div/div[5]/button'
config_element = '//*[@id="nav"]/li[5]/a'
cam_config_element = '//*[@id="nav"]/li[5]/a'
maintenance_element = '//*[@id="menu"]/div/div[2]/div[3]/span'
reboot_button_element = '//*[@id="maintainUpgrade"]/div[1]/div[2]/span[1]/button'
reboot_ok_button_element = '//*[@id="config"]/div[1]/div/table/tbody/tr[2]/td[2]/div/table/tbody/tr[3]/td/div/button[1]'
camera_management_element = '//*[@id="menu"]/div/div[2]/div[5]/span'
#base row
table_row = '//*[@id="tableDigitalChannels"]/div/div[2]/div/'
#/span[3] -> camera name
camera_name_element = '//*[@id="tableDigitalChannels"]/div/div[2]/div[1]/span[3]'
#/span[4] -> camera ip
camera_ip_element = '//*[@id="tableDigitalChannels"]/div/div[2]/div[1]/span[4]'
#/span[8] -> camera status
camera_status_element = '//*[@id="tableDigitalChannels"]/div/div[2]/div[1]/span[8]'
timeout = 5

username_input_element = '//*[@id="portal"]/div/div/div/div[2]/div/div/form/div[1]/div/div/input'
password_input_element = '//*[@id="portal"]/div/div/div/div[2]/div/div/form/div[2]/div/div/input'
login_200_button_element = '//*[@id="portal"]/div/div/div/div[2]/div/div/form/div[4]/div/button'

nvr_username = 'admin'
nvr_password = 'uc-1enviar'

#open nvr 198
def open_admin_198():
   nvr_ip = 'http://172.21.0.198/'
   try:
      chrome_options = Options()
      #chrome_options.add_argument("--incognito")
      chrome_options.add_experimental_option("detach", True)
      global driver_198
      driver_198 = webdriver.Chrome(options=chrome_options)
      #set position and size
      driver_198.set_window_size(1152, 648)
      driver_198.get(nvr_ip)
      login_198()
      #run 24/7
      while True:
         scan()
         #refresh page
         driver_198.refresh()
         time.sleep(timeout)
         #check if viewer got logged out
         if driver_198.current_url != 'http://172.21.0.198/doc/page/config.asp':
           print('logged out...', True)
           login_198()
         #wait 2:00 sesc before refresh page
         time.sleep(120)
   except Exception as e:
      print('error admin 198',e)
      driver_198.close()
      open_admin_198()

def login_198():
   print('logging in 198...')
   wait_for_element_load(login_button_element,driver_198)
   driver_198.find_element(By.ID,'username').send_keys(nvr_username)
   driver_198.find_element(By.ID,'password').send_keys(nvr_password)
   driver_198.find_element(By.XPATH,login_button_element).click()
   wait_for_element_load(config_element,driver_198)
   driver_198.find_element(By.XPATH, config_element).click()
   wait_for_element_load(camera_management_element,driver_198)
   driver_198.find_element(By.XPATH,camera_management_element).click()
   time.sleep(timeout)

#check table if there's offline
def scan():
   print('scanning...')
   camera_count = 0
   is_tail = False
   while is_tail is False:
      try:
         #click row
         driver_198.find_element(By.XPATH,table_row + str([camera_count+1])).click()
         camera_count = camera_count + 1
         #get name and status of current row
         global camera_name
         camera_name = driver_198.find_element(By.XPATH,table_row + str([camera_count]) + '/span[3]').text
         global camera_ip
         camera_ip = driver_198.find_element(By.XPATH,table_row + str([camera_count])  + '/span[4]').text
         camera_status = driver_198.find_element(By.XPATH,table_row + str([camera_count]) + '/span[8]').text
         print(camera_name,camera_status)
         #ignore old cams
         if  (camera_ip == '192.168.74.114' or camera_ip == '192.168.65.14' or camera_ip == '192.168.78.113'):
            print('ignore')
         else:
            #if offline, get ip and reboot
            if camera_status != 'Online':
               reboot(camera_ip,camera_name)
               print('reboot')
      except Exception as e:
         is_tail = True
         print(is_tail,e)
   print('standby...')

#access ip and try reboot
def reboot(ip,name):
   try:
      chrome_options = Options()
      #chrome_options.add_argument("--incognito")
      chrome_options.add_experimental_option("detach", True)
      global cam_driver
      cam_driver = webdriver.Chrome(options=chrome_options)
      #set position and size
      cam_driver.set_window_size(1280, 720)
      #cam_driver.set_window_position(350, 20, windowHandle='current')
      cam_driver.get('http://'+ip)
      success = True
      #credentials of selected ip cam
      username = 'admin'
      password1 = '123456@ad'
      password2 = 'scores135792468'
      time.sleep(timeout)
      wait_for_element_load(login_button_element,cam_driver)
      pre_login_url = cam_driver.current_url
      #login using first password
      cam_driver.find_element(By.ID,'username').send_keys(username)
      cam_driver.find_element(By.ID,'password').send_keys(password1)
      cam_driver.find_element(By.XPATH,login_button_element).click()
      time.sleep(timeout)
      post_login_url = cam_driver.current_url
      #check if password is error
      if pre_login_url == post_login_url:
         cam_driver.find_element(By.ID,'username').send_keys(username)
         cam_driver.find_element(By.ID,'password').send_keys(password2)
         cam_driver.find_element(By.XPATH,login_button_element).click()
      time.sleep(timeout)
      wait_for_element_load(cam_config_element,cam_driver)
      cam_driver.find_element(By.XPATH, cam_config_element).click()
      wait_for_element_load(maintenance_element,cam_driver)
      cam_driver.find_element(By.XPATH,maintenance_element).click()
      wait_for_element_load(reboot_button_element,cam_driver)
      cam_driver.find_element(By.XPATH,reboot_button_element).click()
      wait_for_element_load(reboot_ok_button_element,cam_driver)
      cam_driver.find_element(By.XPATH,reboot_ok_button_element).click()
      time.sleep(timeout)
   except Exception as e:
      print('reboot camera error',e)
      success = False
   finally:
      #close current browser
      cam_driver.close()
      #write log
      write_excel(ip,name,success,198)
      print('is rebooted?..'+name,success)

#function to ensure web element loaded
def wait_for_element_load(element, driver):
   element_wait_load = EC.presence_of_element_located((By.XPATH, element))
   WebDriverWait(driver, 10).until(element_wait_load)
   time.sleep(timeout)

#write logs
def write_excel(ip,name,status,nvr):
   file_name = 'logs.xlsx'
   new_row = [ctime(),ip,name,status,nvr]
   is_file_exist = os.path.isfile(file_name)
   if is_file_exist:
      wb = load_workbook(file_name)
      ws = wb.active
      ws.append(new_row)
   else:
      wb = Workbook()
      ws = wb.active
      ws.append(['Date', 'Ip', 'Name','Rebooted','NVR'])
      ws.append(new_row)
   wb.save(file_name)

open_admin_198()
